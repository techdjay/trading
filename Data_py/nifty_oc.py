from xlrd import open_workbook
import openpyxl
import pymysql
import sqlalchemy
import pandas as pd
import xlrd
import threading
from datetime import datetime
import _thread
import os
#os.ch.dir("C:\Users\Jay\Desktop")
engine = sqlalchemy.create_engine('mysql+pymysql://root:Jayvns9807#@127.0.0.1:3306/option_chain')

def getOptionsData():
    path = "C:\\Users\\Jay\\Desktop\\TXT3LIvedata.xlsx"

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column

    # Loop will print all columns name
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=1, column=i)
        print(cell_obj.value)